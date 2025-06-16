"""
PDF到Excel转换 - 安卓APP版本
使用Kivy框架开发，可编译为APK文件
"""

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.popup import Popup
from kivy.uix.progressbar import ProgressBar
from kivy.clock import Clock
from kivy.utils import platform
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import tempfile
import threading

# 安卓权限请求
if platform == 'android':
    try:
        from android.permissions import request_permissions, Permission
        from android.storage import primary_external_storage_path
    except ImportError:
        # 如果导入失败，定义空函数避免错误
        def request_permissions(permissions):
            pass
        def primary_external_storage_path():
            return '/storage/emulated/0'
        Permission = type('Permission', (), {
            'READ_EXTERNAL_STORAGE': 'android.permission.READ_EXTERNAL_STORAGE',
            'WRITE_EXTERNAL_STORAGE': 'android.permission.WRITE_EXTERNAL_STORAGE',
            'MANAGE_EXTERNAL_STORAGE': 'android.permission.MANAGE_EXTERNAL_STORAGE'
        })

class PDFToExcelApp(App):
    def build(self):
        # 请求安卓权限
        if platform == 'android':
            self.request_android_permissions()

        # 主布局
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # 标题
        title = Label(
            text='PDF到Excel转换工具',
            size_hint=(1, 0.1),
            font_size='20sp'
        )
        main_layout.add_widget(title)
        
        # 文件选择器
        self.file_chooser = FileChooserIconView(
            filters=['*.pdf'],
            size_hint=(1, 0.6)
        )
        main_layout.add_widget(self.file_chooser)
        
        # 选中文件显示
        self.selected_file_label = Label(
            text='未选择文件',
            size_hint=(1, 0.1),
            text_size=(None, None)
        )
        main_layout.add_widget(self.selected_file_label)
        
        # 进度条
        self.progress_bar = ProgressBar(
            max=100,
            value=0,
            size_hint=(1, 0.05)
        )
        main_layout.add_widget(self.progress_bar)
        
        # 状态标签
        self.status_label = Label(
            text='请选择PDF文件',
            size_hint=(1, 0.1)
        )
        main_layout.add_widget(self.status_label)
        
        # 按钮布局
        button_layout = BoxLayout(orientation='horizontal', size_hint=(1, 0.15), spacing=10)

        # 选择文件按钮
        select_btn = Button(text='选择文件')
        select_btn.bind(on_press=self.select_file)
        button_layout.add_widget(select_btn)

        # 开始转换按钮
        self.convert_btn = Button(text='开始转换', disabled=True)
        self.convert_btn.bind(on_press=self.start_conversion)
        button_layout.add_widget(self.convert_btn)

        # 帮助按钮
        help_btn = Button(text='帮助', size_hint=(0.5, 1))
        help_btn.bind(on_press=lambda x: self.show_info_popup())
        button_layout.add_widget(help_btn)

        main_layout.add_widget(button_layout)
        
        return main_layout

    def request_android_permissions(self):
        """请求安卓权限"""
        try:
            permissions = [
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE,
            ]
            # Android 11+ 需要 MANAGE_EXTERNAL_STORAGE 权限
            if hasattr(Permission, 'MANAGE_EXTERNAL_STORAGE'):
                permissions.append(Permission.MANAGE_EXTERNAL_STORAGE)

            request_permissions(permissions)
        except Exception as e:
            print(f"权限请求失败: {e}")

    def get_storage_path(self):
        """获取存储路径"""
        if platform == 'android':
            try:
                return primary_external_storage_path()
            except:
                return '/storage/emulated/0'
        else:
            return os.path.expanduser('~')

    def select_file(self, instance):
        """选择文件"""
        if self.file_chooser.selection:
            selected_file = self.file_chooser.selection[0]
            self.selected_file_label.text = f'已选择: {os.path.basename(selected_file)}'
            self.convert_btn.disabled = False
            self.status_label.text = '文件已选择，可以开始转换'
        else:
            self.show_popup('错误', '请选择一个PDF文件')
    
    def start_conversion(self, instance):
        """开始转换过程"""
        if not self.file_chooser.selection:
            self.show_popup('错误', '请先选择PDF文件')
            return
        
        selected_file = self.file_chooser.selection[0]
        
        # 禁用按钮，显示进度
        self.convert_btn.disabled = True
        self.progress_bar.value = 0
        self.status_label.text = '正在处理...'
        
        # 在后台线程中执行转换
        thread = threading.Thread(target=self.convert_file, args=(selected_file,))
        thread.daemon = True
        thread.start()
    
    def convert_file(self, file_path):
        """转换文件的主要逻辑"""
        try:
            # 更新进度
            Clock.schedule_once(lambda dt: self.update_progress(20, '正在读取PDF文件...'), 0)
            
            # 提取PDF数据
            df = self.extract_pdf_to_dataframe(file_path)
            if df is None:
                Clock.schedule_once(lambda dt: self.conversion_failed('PDF数据提取失败'), 0)
                return
            
            Clock.schedule_once(lambda dt: self.update_progress(50, '正在处理数据...'), 0)
            
            # 处理数据
            processed_df = self.process_excel_data(df)
            if processed_df is None:
                Clock.schedule_once(lambda dt: self.conversion_failed('数据处理失败'), 0)
                return
            
            Clock.schedule_once(lambda dt: self.update_progress(80, '正在生成Excel文件...'), 0)
            
            # 保存Excel文件
            output_path = self.save_to_excel(processed_df, file_path)
            if output_path:
                Clock.schedule_once(lambda dt: self.conversion_success(output_path), 0)
            else:
                Clock.schedule_once(lambda dt: self.conversion_failed('Excel文件保存失败'), 0)
                
        except Exception as e:
            Clock.schedule_once(lambda dt: self.conversion_failed(f'转换过程出错: {str(e)}'), 0)
    
    def extract_pdf_to_dataframe(self, pdf_file):
        """从PDF提取数据"""
        try:
            tables = []
            with pdfplumber.open(pdf_file) as pdf:
                for page in pdf.pages:
                    page_tables = page.extract_tables()
                    if page_tables:
                        for table in page_tables:
                            if table:
                                tables.extend(table)
                    else:
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            for line in lines:
                                if line.strip():
                                    row = line.split('\t') if '\t' in line else line.split()
                                    if len(row) > 1:
                                        tables.append(row)
            
            if not tables:
                return None
            
            df = pd.DataFrame(tables)
            df = df.dropna(how='all').dropna(axis=1, how='all')
            df = df.reset_index(drop=True)
            
            return df
        except Exception:
            return None
    
    def process_excel_data(self, df):
        """处理Excel数据"""
        try:
            processed_df = df.copy()
            
            # 确保有足够的列
            if len(processed_df.columns) < 9:
                for i in range(len(processed_df.columns), 9):
                    processed_df[i] = ''
            
            # 删除A、H、I列
            columns_to_drop = []
            if 0 < len(processed_df.columns):
                columns_to_drop.append(0)
            if 7 < len(processed_df.columns):
                columns_to_drop.append(7)
            if 8 < len(processed_df.columns):
                columns_to_drop.append(8)
            
            for col_idx in sorted(columns_to_drop, reverse=True):
                processed_df = processed_df.drop(processed_df.columns[col_idx], axis=1)
            
            processed_df.columns = range(len(processed_df.columns))
            
            # 处理时间格式和排序
            b_column_index = 0
            g_column_index = 5 if len(processed_df.columns) > 5 else len(processed_df.columns) - 1
            
            if g_column_index < len(processed_df.columns) and b_column_index < len(processed_df.columns):
                try:
                    processed_df = processed_df.sort_values(
                        by=[g_column_index, b_column_index], 
                        ascending=[True, False],
                        na_position='last'
                    )
                    
                    def format_time(time_str):
                        try:
                            if pd.isna(time_str) or time_str == '':
                                return time_str
                            dt = pd.to_datetime(time_str, errors='coerce')
                            if pd.isna(dt):
                                return time_str
                            return dt.strftime('%Y-%m-%d %H:%M')
                        except:
                            return time_str
                    
                    processed_df[b_column_index] = processed_df[b_column_index].apply(format_time)
                    
                except:
                    processed_df = processed_df.sort_values(
                        by=[g_column_index, b_column_index], 
                        ascending=[True, False],
                        na_position='last'
                    )
                
                processed_df = processed_df.reset_index(drop=True)
            
            return processed_df
        except Exception:
            return None
    
    def save_to_excel(self, df, original_file_path):
        """保存为Excel文件"""
        try:
            # 生成输出文件路径
            base_name = os.path.splitext(os.path.basename(original_file_path))[0]

            # 优先使用下载目录，如果不存在则使用存储根目录
            storage_path = self.get_storage_path()
            downloads_path = os.path.join(storage_path, 'Download')

            if not os.path.exists(downloads_path):
                # 尝试创建下载目录
                try:
                    os.makedirs(downloads_path, exist_ok=True)
                except:
                    # 如果创建失败，使用存储根目录
                    downloads_path = storage_path
            
            output_path = os.path.join(downloads_path, f"{base_name}_processed.xlsx")
            
            # 处理文件名冲突
            counter = 1
            while os.path.exists(output_path):
                output_path = os.path.join(downloads_path, f"{base_name}_processed_{counter}.xlsx")
                counter += 1
            
            # 保存Excel文件
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                if worksheet.max_row > 0 and worksheet.max_column > 0:
                    # 设置列宽
                    default_width = 8.43
                    a_column_width = default_width * 2
                    f_column_width = default_width * 3
                    
                    if worksheet.max_column >= 1:
                        worksheet.column_dimensions[get_column_letter(1)].width = a_column_width
                        for row in range(1, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=1)
                            cell.alignment = Alignment(horizontal='left')
                    
                    if worksheet.max_column >= 6:
                        worksheet.column_dimensions[get_column_letter(6)].width = f_column_width
            
            return output_path
        except Exception:
            return None
    
    def update_progress(self, value, status):
        """更新进度"""
        self.progress_bar.value = value
        self.status_label.text = status
    
    def conversion_success(self, output_path):
        """转换成功"""
        self.progress_bar.value = 100
        self.status_label.text = '转换完成！'
        self.convert_btn.disabled = False
        self.show_popup('成功', f'文件已保存到:\n{output_path}')
    
    def conversion_failed(self, error_msg):
        """转换失败"""
        self.progress_bar.value = 0
        self.status_label.text = '转换失败'
        self.convert_btn.disabled = False
        self.show_popup('错误', error_msg)
    
    def show_popup(self, title, message):
        """显示弹窗"""
        popup_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # 消息标签，支持自动换行
        message_label = Label(
            text=message,
            text_size=(300, None),
            halign='center',
            valign='middle'
        )
        popup_layout.add_widget(message_label)

        # 按钮布局
        button_layout = BoxLayout(orientation='horizontal', size_hint=(1, 0.3), spacing=10)

        close_btn = Button(text='确定')
        button_layout.add_widget(close_btn)

        popup_layout.add_widget(button_layout)

        popup = Popup(
            title=title,
            content=popup_layout,
            size_hint=(0.9, 0.5),
            auto_dismiss=False
        )

        close_btn.bind(on_press=popup.dismiss)
        popup.open()

    def show_info_popup(self):
        """显示应用信息"""
        info_text = """PDF到Excel转换工具 v1.0

功能：
• 提取PDF文件中的表格数据
• 自动处理和格式化数据
• 导出为Excel格式文件

使用方法：
1. 选择PDF文件
2. 点击开始转换
3. 等待处理完成

文件保存位置：
/storage/emulated/0/Download/

开发者：PDF Tools Team"""

        self.show_popup('关于应用', info_text)

if __name__ == '__main__':
    PDFToExcelApp().run()
